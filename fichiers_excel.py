import openpyxl

wb1 = openpyxl.load_workbook("data/octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("data/novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("data/decembre.xlsx", data_only=True)



def ajout_donnees_depuis_wb(wb, donnees_dict):
    sheet = wb[wb.sheetnames[0]]
    for i in range(2, sheet.max_row+1):
        nom_article = sheet.cell(i, 1).value
        value_article = sheet.cell(i, 4).value
        if not nom_article:
            break
        
        if donnees_dict.get(nom_article):
            donnees_dict[nom_article].append(value_article)
        else:
            donnees_dict[nom_article] = [value_article]


donnees= {}
ajout_donnees_depuis_wb(wb1, donnees)
ajout_donnees_depuis_wb(wb2, donnees)
ajout_donnees_depuis_wb(wb3, donnees)


# Générer un fichier excel
wb_sortie = openpyxl.Workbook()
sheet = wb_sortie.active

sheet["A1"] = "Article"
sheet["B1"] = "octobre"
sheet["C1"] = "novembre"
sheet["D1"]= "decembre"

row = 2
for article, ventes in donnees.items():
    sheet.cell(row, 1).value = article
    column = 2
    for vente in ventes:
        sheet.cell(row, column).value = vente
        column += 1
    row += 1

wb_sortie.save("data/totale_vente_trimestre.xlsx")
